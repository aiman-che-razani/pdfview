import pytest
import pandas as pd
from openpyxl import Workbook
from excel_manager import list_excels, load_excel, save_excel


@pytest.fixture
def setup_test_excel(tmp_path):
    """Creates a temporary Excel file for testing."""
    test_file = tmp_path / "test.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Name", "Age", "City"])
    ws.append(["Alice", 30, "New York"])
    ws.append(["Bob", 25, "Los Angeles"])
    wb.save(test_file)
    return test_file


def test_list_excels(tmp_path):
    """Tests listing Excel files in a directory."""
    (tmp_path / "file1.xlsx").touch()
    (tmp_path / "file2.xls").touch()
    (tmp_path / "not_excel.txt").touch()

    excel_files = list_excels(tmp_path)
    assert sorted(excel_files) == ["file1.xlsx", "file2.xls"]


def test_load_excel(setup_test_excel):
    """Tests loading an Excel workbook."""
    wb = load_excel(str(setup_test_excel))
    assert isinstance(wb, Workbook)
    assert "Sheet1" in wb.sheetnames


def test_save_excel(setup_test_excel):
    """Tests saving data to an Excel file without affecting other sheets."""
    df = pd.DataFrame({"Name": ["Charlie"], "Age": [28], "City": ["Chicago"]})

    # Save to the test Excel file
    save_excel(df, str(setup_test_excel), "Sheet1")

    # Reload and verify changes
    wb = load_excel(str(setup_test_excel))
    sheet = wb["Sheet1"]

    assert sheet.cell(row=1, column=1).value == "Name"
    assert sheet.cell(row=2, column=1).value == "Charlie"
    assert sheet.cell(row=2, column=3).value == "Chicago"

    # Ensure other sheets remain untouched
    assert len(wb.sheetnames) == 1  # Only one sheet should exist
