# excel_manager.py - Handles Excel file management

import streamlit as st
import pandas as pd
import os
from openpyxl import load_workbook, Workbook


def list_excels(folder):
    """Lists all Excel files in the given folder."""
    return [f for f in os.listdir(folder) if f.endswith(('.xls', '.xlsx'))]


def load_excel(file_path):
    """Loads the Excel file and returns the workbook object."""
    return load_workbook(file_path)


def save_excel(df, file_path, sheet_name):
    """Saves only the modified sheet in the Excel file without affecting other sheets."""
    workbook = load_workbook(file_path)

    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        sheet.delete_rows(1, sheet.max_row)  # Clear only this sheet
    else:
        sheet = workbook.create_sheet(title=sheet_name)

    # Write DataFrame back to the selected sheet
    for col_idx, col_name in enumerate(df.columns, start=1):
        sheet.cell(row=1, column=col_idx, value=col_name)

    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=value)

    workbook.save(file_path)
    workbook.close()  # Close the file safely


def excel_manager_ui(folder):
    """Streamlit UI for managing Excel files."""
    st.header("Excel Manager")

    if not os.path.exists(folder):
        st.error(f"Folder '{folder}' not found.")
        return

    excel_files = list_excels(folder)

    if "selected_excel" not in st.session_state:
        st.session_state.selected_excel = excel_files[0] if excel_files else None
    if "selected_sheet" not in st.session_state:
        st.session_state.selected_sheet = None

    # File Selection
    selected_excel = st.selectbox("Select an Excel file", excel_files)
    st.session_state.selected_excel = selected_excel
    file_path = os.path.join(folder, selected_excel)

    # Handle Excel reload when necessary
    if "reload_excel" not in st.session_state:
        st.session_state["reload_excel"] = False

    if st.session_state["reload_excel"]:
        st.session_state["reload_excel"] = False
        st.rerun()  # Full refresh to update viewer

    # Load Workbook and Sheets
    workbook = load_excel(file_path)
    sheet_names = workbook.sheetnames

    # **Sheet Management**
    st.subheader("Sheet Management")
    col_add, col_delete, col_rename = st.columns(3)

    with col_add:
        new_sheet_name = st.text_input("New Sheet Name:", key="new_sheet")
        if st.button("Add Sheet"):
            if new_sheet_name and new_sheet_name not in workbook.sheetnames:
                workbook.create_sheet(title=new_sheet_name)
                workbook.save(file_path)
                st.rerun()

    with col_delete:
        sheet_to_delete = st.selectbox("Select Sheet to Delete", sheet_names)
        if st.button("Delete Sheet"):
            if len(sheet_names) > 1:
                workbook.remove(workbook[sheet_to_delete])
                workbook.save(file_path)
                st.rerun()
            else:
                st.warning("You cannot delete the last remaining sheet.")

    with col_rename:
        rename_from = st.selectbox("Select Sheet to Rename", sheet_names)
        rename_to = st.text_input("Rename Sheet To:", key="rename_sheet")
        if st.button("Rename Sheet"):
            if rename_to and rename_to not in workbook.sheetnames:
                sheet = workbook[rename_from]
                sheet.title = rename_to
                workbook.save(file_path)
                st.rerun()

    # Sheet Selection
    selected_sheet = st.selectbox("Select a Sheet", sheet_names)
    st.session_state.selected_sheet = selected_sheet
    sheet = workbook[selected_sheet]

    # Extract data from the worksheet
    data = list(sheet.iter_rows(values_only=True))

    if not data:
        df = pd.DataFrame(columns=["Column1"])
    else:
        # Ensure column names are unique
        seen = {}
        new_columns = []
        for col in list(data[0]) if data[0] else ["Column1"]:
            if col in seen:
                seen[col] += 1
                new_columns.append(f"{col} ({seen[col]})")
            else:
                seen[col] = 0
                new_columns.append(col)

        df = pd.DataFrame(data[1:], columns=new_columns)

    edited_df = st.data_editor(df, num_rows="dynamic", key=f"editor_{selected_excel}_{selected_sheet}")

    # **Column Management**
    st.subheader("Column Operations")
    col1, col2, col3 = st.columns(3)
    with col1:
        new_col_name = st.text_input("New Column Name:", key="new_column")
        if st.button("Add Column"):
            if new_col_name and new_col_name not in edited_df.columns:
                edited_df[new_col_name] = ""
                save_excel(edited_df, file_path, selected_sheet)
                st.rerun()

    with col2:
        columns_to_delete = st.multiselect("Select Columns to Delete", edited_df.columns)
        if st.button("Delete Selected Columns"):
            edited_df.drop(columns=columns_to_delete, inplace=True)
            save_excel(edited_df, file_path, selected_sheet)
            st.rerun()

    with col3:
        rename_col = st.selectbox("Select Column to Rename", edited_df.columns)
        rename_col_new = st.text_input("Rename Column To:", key="rename_column")
        if st.button("Rename Column"):
            edited_df.rename(columns={rename_col: rename_col_new}, inplace=True)
            save_excel(edited_df, file_path, selected_sheet)
            st.rerun()

    # **Row Management**
    st.subheader("Row Operations")
    col4, col5 = st.columns(2)
    with col4:
        if st.button("Add Row"):
            edited_df.loc[len(edited_df)] = [""] * len(edited_df.columns)
            save_excel(edited_df, file_path, selected_sheet)
            st.rerun()

    with col5:
        row_to_delete = st.number_input("Enter Row Index to Delete", min_value=0, max_value=max(len(edited_df) - 1, 0))
        if st.button("Delete Selected Row"):
            edited_df.drop(index=row_to_delete, inplace=True)
            edited_df.reset_index(drop=True, inplace=True)
            save_excel(edited_df, file_path, selected_sheet)
            st.rerun()

    # **Push Column Values to First Row**
    st.subheader("Push Column Values to First Row")
    if st.button("Push Columns to First Row"):
        if not df.empty:
            # Copy first row values
            first_row_values = df.iloc[0].copy()
            new_row = pd.DataFrame([first_row_values], columns=df.columns)

            # Move headers to first row and insert new row
            df.iloc[0] = df.columns
            df = pd.concat([df.iloc[:1], new_row, df.iloc[1:]]).reset_index(drop=True)

            # Rename column names to "Column X" (X = count of non-empty column names)
            non_empty_columns = sum(1 for col in df.columns if col.strip())  # Count non-empty column names
            df.columns = [f"Column {i + 1}" for i in range(non_empty_columns)]  # Rename dynamically

            # Save back to Excel
            save_excel(df, file_path, selected_sheet)
            st.success("Column values pushed and columns renamed to 'Column X'.")
            st.rerun()

    # Save Changes
    st.subheader("Save Changes")
    if st.button("Save Changes"):
        save_excel(edited_df, file_path, selected_sheet)
        st.success("Changes saved successfully!")
